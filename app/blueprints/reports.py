from flask import Blueprint, render_template, Response, send_file, flash, redirect, url_for
from flask_login import login_required
from app.models import Product, Order, User, Vendor
from app.decorators import role_required
import csv
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime

reports_bp = Blueprint('reports', __name__)

@reports_bp.route('/reports')
@login_required
@role_required('super_admin', 'admin', 'manager')
def index():
    return render_template('reports/index.html', title='Reports & Analytics')

@reports_bp.route('/reports/export/<type>')
@login_required
def export_csv(type):
    """Export reports as Excel files with formatting"""
    wb = Workbook()
    ws = wb.active
    
    # Header styling
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    if type == 'inventory':
        ws.title = "Inventory Report"
        headers = ['Code', 'Name', 'Category', 'Brand', 'Stock', 'Unit', 'Cost Price', 'Selling Price', 'Stock Value']
        ws.append(headers)
        
        # Style headers
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        products = Product.query.all()
        for p in products:
            stock_value = float(p.cost_price or 0) * (p.stock_quantity or 0)
            ws.append([
                p.code, p.name, p.category, p.brand,
                p.stock_quantity, p.unit,
                float(p.cost_price or 0), float(p.selling_price or 0),
                stock_value
            ])
        
        # Add total row
        last_row = len(products) + 2
        ws[f'E{last_row}'] = 'TOTAL'
        ws[f'E{last_row}'].font = Font(bold=True)
        ws[f'I{last_row}'] = f'=SUM(I2:I{last_row-1})'
        ws[f'I{last_row}'].font = Font(bold=True)
        
        filename = f'inventory_report_{datetime.now().strftime("%Y%m%d")}.xlsx'
        
    elif type == 'purchase':
        ws.title = "Purchase Report"
        headers = ['Order ID', 'Date', 'Vendor', 'Total Amount', 'Status']
        ws.append(headers)
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
        
        purchases = Order.query.filter_by(type='PURCHASE').all()
        for p in purchases:
            v_name = p.vendor.name if p.vendor else 'N/A'
            ws.append([
                p.id,
                p.date.strftime('%Y-%m-%d') if p.date else '',
                v_name,
                float(p.total_amount or 0),
                p.status
            ])
        
        # Add total
        last_row = len(purchases) + 2
        ws[f'C{last_row}'] = 'TOTAL'
        ws[f'C{last_row}'].font = Font(bold=True)
        ws[f'D{last_row}'] = f'=SUM(D2:D{last_row-1})'
        ws[f'D{last_row}'].font = Font(bold=True)
        
        filename = f'purchase_report_{datetime.now().strftime("%Y%m%d")}.xlsx'

    elif type == 'vendors':
        ws.title = "Vendor Directory"
        headers = ['Vendor Name', 'Contact Person', 'Phone', 'Email', 'GSTIN', 'Address']
        ws.append(headers)
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
        
        vendors = Vendor.query.all()
        for v in vendors:
            ws.append([
                v.name,
                v.contact_person or '',
                v.phone or '',
                v.email or '',
                v.gstin or '',
                v.address or ''
            ])
        
        filename = f'vendor_directory_{datetime.now().strftime("%Y%m%d")}.xlsx'
        
    elif type == 'sales':
        ws.title = "Sales Report"
        headers = ['Order ID', 'Date', 'Customer', 'Total Amount', 'Tax', 'Grand Total', 'Status']
        ws.append(headers)
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            
        sales = Order.query.filter_by(type='SALE').all()
        for s in sales:
            c_name = s.customer.name if s.customer else 'Guest'
            ws.append([
                s.id,
                s.date.strftime('%Y-%m-%d') if s.date else '',
                c_name,
                float(s.total_amount or 0),
                float(s.tax_amount or 0),
                float(s.grand_total or 0),
                s.status
            ])
            
        # Add totals
        last_row = len(sales) + 2
        ws[f'C{last_row}'] = 'TOTALS'
        ws[f'C{last_row}'].font = Font(bold=True)
        ws[f'D{last_row}'] = f'=SUM(D2:D{last_row-1})'
        ws[f'F{last_row}'] = f'=SUM(F2:F{last_row-1})'
        ws[f'D{last_row}'].font = Font(bold=True)
        ws[f'F{last_row}'].font = Font(bold=True)
        
        filename = f'sales_report_{datetime.now().strftime("%Y%m%d")}.xlsx'
        
    else:
        return "Invalid Report Type", 400
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )

@reports_bp.route('/reports/send-daily-summary')
@login_required
@role_required('super_admin', 'admin')
def send_daily_summary():
    """Manually trigger daily summary email"""
    try:
        from app.email_service import EmailService
        EmailService.send_daily_summary()
        flash('Daily summary email sent successfully!', 'success')
    except Exception as e:
        flash(f'Failed to send email: {str(e)}', 'danger')
    return redirect(url_for('reports.index'))

@reports_bp.route('/reports/test-email')
@login_required
@role_required('super_admin', 'admin')
def test_email():
    """Send a test email"""
    try:
        from app.email_service import EmailService
        from app.models import Product
        
        # Send low stock alert as test
        low_stock_products = Product.query.filter(
            Product.stock_quantity <= Product.min_stock_alert
        ).all()
        
        if low_stock_products:
            EmailService.send_low_stock_alert(low_stock_products)
            flash(f'Test email sent! Low stock alert for {len(low_stock_products)} products.', 'success')
        else:
            flash('No low stock products to test email with. All stock levels are good!', 'info')
    except Exception as e:
        flash(f'Test email failed: {str(e)}', 'danger')
    
    return redirect(url_for('reports.index'))


